#!/usr/bin/env python
# -*- coding=utf-8 -*-

import os
import argparse
# import codecs
import openpyxl
import gzip
from functools import partial

def mkdir_if_not_exists(path):
    if not os.path.exists(path):
        # print 'make a new dir:', path
        # os.makedirs(path, mode=0o777, exist_ok=False)
        os.makedirs(path)
def safe_open(filename, mode='r'):

    try:
        if mode == 'w':
            dirname = os.path.dirname(filename)
            if dirname and not os.path.exists(dirname):
                os.makedirs(dirname)
        if filename.endswith('.gz'):
            mode += 'b'
            return gzip.open(filename, mode)

        return open(filename, mode)
    except Exception as e:
        print('[error] fail to open file:%s' % filename)
        exit(1)

def get_sheetname(filename,sheetnames=None,sheetname_idx=[0]):

    filename = os.path.basename(filename).split('.')

    if not sheetnames:
        sheetname = ".".join([filename[idx] for idx in sheetname_idx])
    else:
        sheetname = sheetnames.pop(0)

    return sheetname[:31]


def text2excel(outfile, *infiles):
    '''
    infiles can be a string: 'a.xls,b.xls'          ('a.xls,b.xls', )
                  or a list: ['a.xls', 'b.xls']     (['a.xls', 'b.xls'], )
    or many positional args: 'a.xls', 'b.xls'       ('a.xls', 'b.xls')
    '''

    # wb = openpyxl.Workbook(encoding='utf8')
    wb = openpyxl.Workbook()

    if len(infiles) == 1:
        if isinstance(infiles[0], str):
            infile_list = infiles[0].split(',')
        elif isinstance(infiles[0], list):
            infile_list = infiles[0]

    elif len(infiles) >= 2:
        infile_list = list(infiles)

    else:
        exit('error infiles: {}'.format(infiles))

    for infile in infile_list:
        sheetname = get_sheetname(filename=infile)
        print('create sheet:%s' % sheetname)
        sheet = wb.create_sheet(title=sheetname)
        # with codecs.open(infile, mode='r', encoding='gbk', errors='ignore') as f:
        with safe_open(infile) as f:
            for n, line in enumerate(f):
                row = n + 1
                linelist = line.strip().split('\t')
                for m, value in enumerate(linelist):
                    column = m + 1
                    sheet.cell(row=row, column=column, value=value)

    # remove default sheet
    try:
        wb.remove(wb['Sheet'])
    except AttributeError:
        wb.remove_sheet(wb.get_sheet_by_name('Sheet'))  # for the old version

    outdir = os.path.dirname(outfile)
    if outdir:
        mkdir_if_not_exists(outdir)

    wb.save(filename=outfile)

    print('write excel file:%s' % outfile)


if __name__ == "__main__":
    usage = '''usage:
    text2excel.py <outfile> <--sheetname|--sheetname_idx> <infiles>
        outfile   the output excel format file
        infiles   the input text format files, separate by comma
    '''
    parser = argparse.ArgumentParser(
        formatter_class = argparse.RawDescriptionHelpFormatter,
        description = usage,
    )
    parser.add_argument('ofile',help="Output excel file name")
    parser.add_argument('--sheetname','-n',help="alist of sheetname seperated by commas")
    parser.add_argument('--sheetname_idx','-i',default="0",help="you can use part of input filename as\
output sheetname which split by dot,for example 0,1,2 sheetname of excel will use first three part of input filename ")
    parser.add_argument('infile',nargs="+",help="alist of infile required to merge to a excel")
    args = vars(parser.parse_args())
    if args.get('sheetname'):
        tags = args.get("sheetname").split(',')
        get_sheetname = partial(get_sheetname,sheetnames=tags)
    else:
        tags = [int(i) for i in args.get('sheetname_idx').split(',')]
        get_sheetname = partial(get_sheetname,sheetname_idx=tags)

    text2excel(args.get("ofile") ,*args.get("infile"))

