#!/usr/bin/env python

# -*- coding=utf-8 -*-

#Extracting indicated sheet to plain text

import argparse
import datetime
from openpyxl import load_workbook


time = datetime.datetime.now().strftime('%Y-%m-%d')

def excel_to_txt(args):
    excel = args.get('excel')
    sheet_names = args.get('sheet_name')
    sheet_numbers = args.get('sheet_number') 
    book = load_workbook(excel)
    if args.get('show_sheets'):
        print('All sheets in the excel lists blow:\n')
        for idx,sheetname in enumerate(book.sheetnames):
            print('[{idx}] {sheetname}'.format(idx=idx,sheetname=sheetname))
    else:
        sheet_list = []
        if sheet_numbers:
            sheet_list += [book.sheetnames[i] for i in sheet_numbers]
        if sheet_names and (sheet_names not in sheet_list):
            sheet_list += sheet_names
        if (not sheet_numbers) and (not sheet_names):
            sheet_list = book.sheetnames

        for sheet in sheet_list:
            ofile = sheet + '.' + time + '.xls'
            print('Writing %s to %s' % (sheet,ofile))
            with open(ofile,'w') as odata:
                for row in book[sheet]:
                    row_list = [cell.value for cell in row]
                    row_list = [args.get('handle_NA') if i is None else str(i) for i in row_list]
                    odata.write("\t".join(row_list) + "\n")
        print('Writing Completion.') 

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
        Example:
        
            excel_to_txt.py -e excel  # extract all sheets 
            excel_to_txt.py -e excel --sheet_number 0 1 2
            excel_to_txt.py -e excel --sheet_name "sheet1" "sheetx"
        '''
    )
    parser.add_argument('--excel','-e',required=True,help='Excel Name')
    parser.add_argument('--sheet_number','-num',nargs="*",default=[],type=int,help='The index of sheet number zero-based')
    parser.add_argument('--sheet_name','-name',nargs="*",default=[],help='One or more sheet names')
    parser.add_argument('--handle_NA','-NA',default='',
        help='what shold be used to replace None objects[default:""]')
    parser.add_argument('--show_sheets','-show',action='store_true',help='Show the sheets of excel instead of extracting them')
    args = parser.parse_args()
    excel_to_txt(vars(args))
