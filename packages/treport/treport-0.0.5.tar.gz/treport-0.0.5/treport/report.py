import configparser
import datetime
import os.path
from sys import platform
from pathlib import Path
from lxml import etree
import openpyxl as xlsx
import postgres as pg
from jinja2 import Template, Environment, FileSystemLoader, meta
import logging
from treport.logger import get_logger

DIR = Path(__file__).resolve().parent
XSD_FILE_NAME = os.path.join(DIR, 'treport.xsd')

def get_config(path_to_inifile):
    config = configparser.ConfigParser()
    config.read(path_to_inifile)
    login = config.get('database', 'login')
    password = config.get('database', 'password')
    host = config.get('database', 'host')
    port = config.get('database', 'port')
    database = config.get('database', 'database')
    db_url = f'postgresql://{login}:{password}@{host}:{port}/{database}'
    path_to_params_reports_file = config.get('report', 'params_reports')
    return db_url, path_to_params_reports_file

def validate_params_report(xml_doc, xsd_doc):
    xmlschema = etree.XMLSchema(xsd_doc)
    return xmlschema.validate(xml_doc)

class Report():
    # Код отчета
    codeReport = ''
    # Наименование отчета
    nameReport = ''
    # Путь к файлу шаблона
    template = ''
    # Путь к каталогу, в котором может быть сохранен файл отчета
    outDir = ''
    # Параметры отчета
    params_report = {}

    # Результат проверки XM-файла на соответствие XSD-схеме
    xml_validation_result : bool = None

    # Имя файла отчета
    report_file_name = ''

    '''reportPages - атрибут, в котором хранятся сведения о листе отчета. Атрибут является словарем и имеет следующие ключи:
        pageName - имя листа в xlsx-файле
        sqlFile - путь к файлу, в котором хранится SQL-запрос
        headerRows - количество строк на листе, которые выделены под заголовок
        ignoredColumns - список номеров колонок, которые необходимо игнорировать при формировании отчета
        sqlText - текст SQL-запроса, который сформирован на основании текста в файле с подставленными значениями параметров запроса
        sqlResult - результат выполнения SQL-запроса
        '''
    reportPages = {}
    # Контент сформированного отчета
    contentReport = None

    # Значения с параметрами отчета
    paramValues = {}

    # URL  подключения к БД
    dbUrl = ''

    checkParamsResult : bool = False
    isCorrect : bool = None

    logger = logging.Logger


    def __init__(self, report_code, path_to_params_reports_file, param_values, db_url):
        '''

        :param report_code:  Код отчета
        :param path_to_params_reports_file: Путь к XML-файлу, в котором хранится описание свойств отчетов.
        :param param_values: Значение параметров.
        :param db_url: URL подклбчения к базе данных
        '''
        self.logger = get_logger(__name__)
        self.codeReport = report_code
        self.paramValues = param_values
        self.get_params_report(path_to_params_reports_file)
        if self.isCorrect:
            self.checkParamsResult = self.check_params()
            if self.checkParamsResult:
                self.generate_sql()
                if self.isCorrect:
                    self.dbUrl = db_url
                    self.get_dataset()
                    self.generate_report()
            else:
                self.logger.error(f'Параметры отчета не соответствуют списку параметроа в XML-файле {path_to_params_reports_file}')
        else:
            self.logger.error(f'Проверка XML-файла {path_to_params_reports_file} проведена, файл имеет ошибки')


    def set_iscorrect_true(self):
        if self.isCorrect or self.isCorrect is None:
            self.isCorrect = True

    def set_iscorrect_false(self):
        if self.isCorrect or self.isCorrect is None:
            self.isCorrect = False

    def generate_file_name(self, filename_rule) -> str:
        '''
        Формирует имя файла по правилу, которое описано внутри элемента fileNameGenerateRule.

        :param filename_rule: содержимое элемента fileNameGenerateRule
        :return: имя файла
        '''
        tmp_rule = []
        for item_rule in filename_rule:
            if item_rule.attrib['segmentType'] == 'segmentString':
                tmp_rule.append(item_rule.text)
            if item_rule.attrib['segmentType'] == 'parametr':
                tmp_rule.append(self.paramValues[item_rule.text])
            if item_rule.attrib['segmentType'] == 'function':
                name_function : str = item_rule.text
                if name_function.startswith('now'):
                    l = len(name_function)
                    parametr_func = name_function[5:len(name_function)-2]
                    tmp_rule.append(datetime.datetime.now().strftime(parametr_func))
                else:
                    tmp_rule.append(f'func_{name_function}_not_found')
                    self.logger.warning(f'Функция {name_function} в списке доступных функций не найдена')

        result_file_name = "".join(tmp_rule)
        self.logger.info('Имя файла '+result_file_name+' сформировано')
        return result_file_name

    def get_params_report(self, path_to_params_reports_file):
        _code_report_find = False
        xsd_f = open(XSD_FILE_NAME)
        xml_f = open(path_to_params_reports_file, encoding='utf-8')
        xsd_doc = etree.parse(xsd_f)
        xml_doc = etree.parse(xml_f)
        validation_result = validate_params_report(xml_doc=xml_doc, xsd_doc=xsd_doc)
        self.xml_validation_result = validation_result
        if self.xml_validation_result:
            self.logger.info('Проверка XML-файла '+path_to_params_reports_file+' проведена, файл корректен')
            self.set_iscorrect_true()
            xml_root = xml_doc.getroot()[0]

            for item_report in xml_root:
                if item_report.attrib['codeReport'] == self.codeReport:
                    _code_report_find = True
                    self.logger.info(f'Чтение блока параметров отчета {self.codeReport}')

                    self.nameReport = item_report[0].text       # Наименование отчета
                    self.logger.info(f'Отчет {self.nameReport}')

                    self.template = item_report[1].text         # Путь к файлу-шаблону отчета
                    self.logger.info(f'Шаблон {self.template}')

                    self.outDir = item_report[2].text           # Путь к каталогу, где будет сохраняться отчет
                    self.logger.info(f'Путь к каталогу, где будет сохраняться отчет: {self.template}')

                    params = item_report[4]                     # Параметры отчета
                    self.logger.info(f'Параметры отчета {self.codeReport}')
                    for item_params in params:
                        parametr_code = item_params.attrib['id']
                        self.params_report[parametr_code] = {}
                        for parametr in item_params:
                            self.params_report[parametr_code][parametr.tag] = parametr.text
                            self.logger.info(f'Параметр: {parametr_code}, значение: {parametr.text}')

                    file_name_rule = item_report[3]  # Правило конструирования имени файла
                    self.checkParamsResult = self.check_params()
                    if self.checkParamsResult:
                        self.report_file_name = self.generate_file_name(file_name_rule)
                        self.logger.info(f'Имя файла отчета: {self.template}')
                    else:
                        self.logger.error(f'Невозможно сконструировать имя файла отчета: {self.template}')

                    report_pages = item_report[5]               # Свойства каждой страницы отчета
                    self.logger.info(f'Свойства страниц отчета {self.codeReport}')

                    for item_page in report_pages:
                        page_code = item_page.attrib['codePage']
                        self.reportPages[page_code] = {}
                        for parametr_page in item_page:
                            if parametr_page.tag == 'ignoredColumns':
                                igonred_columns = parametr_page.text.split(',')
                                int_ignored_columns = []
                                for item in igonred_columns:
                                    int_ignored_columns.append(int(item))
                                self.reportPages[page_code][parametr_page.tag] = int_ignored_columns

                            if parametr_page.tag == 'headerRows':
                                self.reportPages[page_code][parametr_page.tag] = int(parametr_page.text)

                            if parametr_page.tag not in ('ignoredColumns', 'headerRows'):
                                self.reportPages[page_code][parametr_page.tag] = parametr_page.text
                        self.logger.info(f'Страница {page_code} параметр {parametr_page} значение {parametr_page.text}')
            if _code_report_find:
                self.set_iscorrect_true()
            else:
                self.set_iscorrect_false()
                self.logger.error(f'Код отчета {self.codeReport} отсутствует в списке допустимых значений' )
        else:
            self.isCorrect = False

    def check_params(self):
        '''
        Проверяет значения передаваемых параметров на соответствие списку допустимых значений.
        :return:
        '''
        missing_parametrs = []
        for item_param_value in self.paramValues:
            if item_param_value not in self.params_report:
                missing_parametrs.append(item_param_value)
        if len(missing_parametrs) > 0:
            self.logger.error(f'В запрос передаются параметры, отсутствующие в списке допустимых: {", ".join(missing_parametrs)}')
            self.set_iscorrect_false()
            return False
        else:
            self.set_iscorrect_true()
            return True


    def generate_sql(self):
        '''
        В содержимом файла SQL-запроса заполняет значения параметров. В текстовом файле должна использоваться нотоция,
        принятая в шаблонизаторе Jinja2.

        Результат работы метода сохраняется в атрибуте класса reportPages.

        :return:
        '''

        for report_page in self.reportPages:
            template_file_name = self.reportPages[report_page]['sqlFile']
            env = Environment(loader=FileSystemLoader(''))
            template_source = env.loader.get_source(env, template_file_name)[0]
            parsed_content = env.parse(template_source)

            missing_parametrs_in_template = []
            for parametr_in_template in meta.find_undeclared_variables(parsed_content):
                if parametr_in_template not in self.params_report:
                    missing_parametrs_in_template.append(parametr_in_template)

            if len(missing_parametrs_in_template) == 0:
                f = open(template_file_name)
                template = Template(f.read())
                self.reportPages[report_page]['sqlText'] = template.render(self.paramValues)
                self.logger.info(f'SQL-файл отчета {self.codeReport} для страницы {report_page} сформирован')
                if self.isCorrect is None:
                    self.isCorrect = True
            else:
                if self.isCorrect or self.isCorrect is None:
                    self.isCorrect = False
                self.logger.error(f'В тексте SQL-запроса присутствуют параметры, отсутствующие в списке допустимых: '
                                  f'{", ".join(missing_parametrs_in_template)}')



    def get_dataset(self):
        db = pg.Postgres(url=self.dbUrl)
        self.logger.info(f'Подключение к {self.dbUrl} выполнено')
        for report_page in self.reportPages:
            sql_text = self.reportPages[report_page]['sqlText']
            self.logger.info(f'Запрос к БД для формирования отчета {self.codeReport} для страницы {report_page} стартует')
            data_set_page = db.all(sql_text)
            self.reportPages[report_page]['sqlResult'] = data_set_page
            self.logger.info(f'Запрос к БД для формирования отчета {self.codeReport} для страницы {report_page} выполнен')

    def delete_rows(self, ws, max_data_rows):
        max_sheet_rows = ws.max_row
        rows_to_delete = max_sheet_rows - max_data_rows
        ws.delete_rows(max_data_rows + 1, rows_to_delete)

    def generate_table(self, ws, header, list_page, code_page):
        '''
        Заполнение листа книги MS Excel.

        :param ws: лист рабочей книги MS Excel
        :param header: количество строк в таблице отведенных под заголовок
        :param list_page: результат запроса к БД
        :param name_sheet: наименование листа рабочей книги MS Excel
        :return:
        '''
        counter = 0
        for item in list_page:
            counter += 1
            line_number = header + counter
            col_count = 0
            for field in item:
                col_count += 1
                if col_count not in self.reportPages[code_page]['ignoredColumns']:
                    ws.cell(row=line_number, column=col_count).value = item[col_count - 1]

        self.logger.info(f'Лист {code_page} отчета {self.codeReport} заполнен')

        self.delete_rows(ws, counter + header)

    def generate_report(self):
        '''
        Функция формирования файла в формате OpenXML (xlsx) по результатам SQL-запроса.

        :return:
        '''
        wb = xlsx.load_workbook(filename=self.template, read_only=False, keep_vba=False, data_only=False)
        self.logger.info(f'Формирование файла отчета {self.codeReport} в формате MS Excel начато')

        for report_page in self.reportPages:
            ws = wb[self.reportPages[report_page]['pageName']]
            header = self.reportPages[report_page]['headerRows']
            dataset_page = self.reportPages[report_page]['sqlResult']
            self.generate_table(ws, header, dataset_page, report_page)
        self.contentReport = wb