#! /usr/bin/env python3
# -*- coding: utf-8 -*-
# @Time    : 2022-06-16 10:07
# @Site    :
# @File    : dealFile.py
# @Software: PyCharm
import os

import openpyxl as op

# def save_excel(save_path: str, fileName: str, data: list):
#     """
#     :param save_path: 路径
#     :param fileName: 文件名（不包含后缀）
#     :param data: ex:[{"id": 1, "name": "立智", "price": 100}, {"id": 2, "name": "维纳", "price": 200},{"id": 3, "name": "如家", "price": 300}]
#     :return:
#     """
#     fileName = f"{fileName}.xlsx"
#     workbook = xw.Workbook(f'{save_path}/{fileName}')  # 创建工作簿
#     bold = workbook.add_format({
#         'bold': True,  # 字体加粗
#         # 'border': 1,  # 单元格边框宽度
#         'align': 'left',  # 水平对齐方式
#         'valign': 'vcenter',  # 垂直对齐方式
#         'text_wrap': True,  # 是否自动换行
#         })
#     worksheet1 = workbook.add_worksheet("sheet1")  # 创建子表
#     worksheet1.activate()  # 激活表
#     title = data[0].keys()  # 设置表头
#     worksheet1.write_row('A1', title, bold)  # 从A1单元格开始写入表头
#     i = 2  # 从第二行开始写入数据
#     for j in range(len(data)):
#         value = list(data[j].values())
#         print([value[i] for i in range(len(value))])
#         insertData = [value[i] for i in range(len(value))]
#         row = 'A' + str(i)
#         worksheet1.write_row(row, insertData, bold)
#         i += 1
#     workbook.close()  # 关闭表

def save_excel(save_path: str, fileName: str, data: list,sheet_name='Sheet',append=False):
    """
    :param save_path: 路径
    :param fileName: 文件名（不包含后缀）
    :param data: ex:[{"id": 1, "name": "立智", "price": 100}, {"id": 2, "name": "维纳", "price": 200},{"id": 3, "name": "如家", "price": 300}]
    :param sheet_name: sheet名
    :param append: 追加写入
    :return:
    """

    if append:
        fileName = f"{save_path}/{fileName}.xlsx"
        assert os.path.isfile(fileName), f"{fileName} 文件不存在"
        wb = op.load_workbook(fileName)  # 创建工作簿对象
        ws = wb[sheet_name]  # 创建子表
        append_sheet = wb.copy_worksheet(ws)
        rows = append_sheet.max_row
        for row in range(rows + 1, rows + len(data) + 1):
            i = 1
            for l in range(1, len(data[0])+1):
                ws.cell(row, l, list(data[i].values())[l-1])
            i += 1
    else:
        fileName = f"{save_path}/{fileName}.xlsx"
        wb = op.Workbook()  # 创建工作簿对象
        ws = wb['Sheet']  # 创建子表
        title = list(data[0].keys())  # 设置表头
        ws.append(title)  # 添加表头
        for i in range(len(data)):
            print(i)
            d = list(data[i].values())
            ws.append(d)  # 每次写入一行
    wb.save(fileName)


def read_excel(path: str, sheet: str):
    """读取excel"""
    # 打开工作簿
    workbook = op.load_workbook(path)
    # 获取表单
    sheet = workbook[sheet]
    # 获取最大行数
    max_row = sheet.max_row
    # 获取最大列数
    max_column = sheet.max_column

    lr = tuple([tuple([sheet.cell(row=row, column=column).value for column in range(1, max_column+1)]) for row in range(1, max_row+1)])
    print(lr)


    return lr

if __name__ == '__main__':
    testData = [{"id": 1, "name": "立智", "price": 100},
                {"id": 2, "name": "维纳", "price": 200},
                {"id": 3, "name": "如家", "price": 300},


                {"id": 3, "name": "如家", "price": 300},
                {"id": 2, "name": "维纳", "price": 200},
                {"id": 2, "name": "维纳", "price": 200},
                ]

    fileName = '测试4'

    save_excel('./', fileName, testData,append=True)
    # read_excel('./测试3.xlsx','Sheet')
